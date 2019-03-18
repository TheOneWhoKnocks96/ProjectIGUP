import unicodedata
import sys
import time
import datetime
import calendar
from os import listdir
from os.path import isfile, join
onlyfiles = [f for f in listdir('1861 - 1867') if isfile(join('1861 - 1867', f))]

from openpyxl import load_workbook

print(onlyfiles)


for file in onlyfiles:
    file='1861 - 1867/' + file
    wb=load_workbook(filename=str(file), data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)



    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                cell413 = ws.cell(row=4, column=13).value
                newcell413 = cell413[-4:]
                anoDez = int(newcell413) - 1
                
            if(ws==wb.worksheets[1]):
                cellyear = str(ws.cell(row=4, column=13).value)
                newcellyear = cellyear[-4:]
                year = newcellyear
                ano_ver=int(year)
            temp = ws.cell(row=4, column=13).value
            temp1 = temp[-4:]
            ano = int(temp1)
            
        #FEVEREIRO
        if ws==wb.worksheets[2]:
            if(calendar.isleap(ano_ver)):
                minr=22
                maxc=32
                maxr=50
                minc=2
            else:
                minr=22
                maxc=32
                maxr=49
                minc=2
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[1] or ws==wb.worksheets[3] or ws==wb.worksheets[5] or ws==wb.worksheets[7] or ws==wb.worksheets[8] or ws==wb.worksheets[10]:
            minr=22
            maxc=32
            maxr=52
            minc=2
           
           
        #Mes com 30 dias    
        if ws==wb.worksheets[4] or ws==wb.worksheets[6] or ws==wb.worksheets[9] or ws==wb.worksheets[11]:
            minr=22
            maxc=32
            maxr=51
            minc=2
            
        if index3 <= 11:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc, values_only=True)):
                for index, cell in enumerate(row):
                    #if(index3==0):
                     #   if(anoDez!=(ano - 1)):
                      #      print('Ano de Dezembro esta mal. ' + file)
                    
                    if(index3!=0):
                        anoValidation = ws.cell(row=4, column=13).value
                        newAnoValidation = anoValidation[-4:]
                        AnoVali = int(newAnoValidation)
                        if (ano!= AnoVali):
                            print('Ano esta mal! ' + file)
                    dia=index2 + 1
                    if(index3==0):
                        mes=12
                    mes=index3 
                    
                    #barometro
                    if(index<31 and index>=0):
                        if(index==1 or index==5 or index==6 or index==14 or index==17):
                            hora = 9
                            try:
                                value=float(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                        if(index==2 or index==7 or index==8 or index==15 or index==18 or index==25 or index==26 or index==27):
                            hora = 12
                            try:
                                value=float(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        
                        if(index==3 or index==9 or index==10 or index==11 or index==12 or index==16 or index==19 or index==20):
                            hora = 15
                            try:
                                value=float(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                            
    
                        if(index==4 or index==13):
                            try:
                                value=float(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        if(index==22 or index==28): 
                            hora = 9
                            try: 
                                value = str(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut


                        if(index==23 or index==29):
                            hora = 12
                            try: 
                                value = str(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        if(index==24 or index==30):
                            hora = 15
                            try: 
                                value = str(cell)
                            except ValueError:
                                print('Valores Errados:')
                                print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia')
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                            
    print(file)                    
    print('----------------------------') 
                            
